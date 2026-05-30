#!/usr/bin/env python3
"""
seed_from_excel.py — Nạp dữ liệu master (Customer/Product/Spec) từ Excel/CSV
thẳng vào CSDL SQLite của MES (ccl_mes.db) để khởi tạo nhanh dữ liệu thực tế.

ETL đơn giản: đọc file -> chuẩn hóa -> UPSERT vào các bảng do EF Core tạo.
Dùng cho giai đoạn dev/SQLite. Khi sang SQL Server, đổi phần kết nối sang pyodbc.

Định dạng cột (Excel .xlsx hoặc .csv) — sheet/header:
    customer_code, customer_name, product_code, product_name, spec_code, spec_title

Cách dùng:
    # nạp từ CSV (không cần thư viện ngoài)
    python3 tools/seed_from_excel.py tools/sample_master.csv --db src/CCL.MES.Web/ccl_mes.db

    # nạp từ Excel (cần: pip install openpyxl)
    python3 tools/seed_from_excel.py master.xlsx --db src/CCL.MES.Web/ccl_mes.db
"""
import argparse
import csv
import os
import sqlite3
import sys
from datetime import datetime, timezone


def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            sys.exit("❌ Cần cài openpyxl để đọc Excel:  pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]
    else:
        with open(path, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))


def now_iso():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def upsert(path, db):
    if not os.path.exists(db):
        sys.exit(f"❌ Không tìm thấy DB {db}. Hãy chạy ứng dụng .NET 1 lần để EF Core tạo schema, rồi chạy lại.")
    rows = read_rows(path)
    if not rows:
        sys.exit("❌ File không có dữ liệu.")

    conn = sqlite3.connect(db)
    conn.execute("PRAGMA foreign_keys = ON;")
    cur = conn.cursor()
    ts = now_iso()
    n_cust = n_prod = n_spec = 0

    def get_or_create_customer(code, name):
        nonlocal n_cust
        cur.execute('SELECT Id FROM Customers WHERE Code=?', (code,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute('INSERT INTO Customers (Code,Name,CreatedAt,CreatedBy) VALUES (?,?,?,?)',
                    (code, name, ts, "seed"))
        n_cust += 1
        return cur.lastrowid

    def get_or_create_product(code, name, customer_id):
        nonlocal n_prod
        cur.execute('SELECT Id FROM Products WHERE ProductCode=?', (code,))
        r = cur.fetchone()
        if r:
            return r[0]
        cur.execute('INSERT INTO Products (ProductCode,Name,CustomerId,CreatedAt,CreatedBy) VALUES (?,?,?,?,?)',
                    (code, name, customer_id, ts, "seed"))
        n_prod += 1
        return cur.lastrowid

    def ensure_spec(code, title, product_id):
        nonlocal n_spec
        cur.execute('SELECT Id FROM Specs WHERE SpecCode=?', (code,))
        if cur.fetchone():
            return
        cur.execute('INSERT INTO Specs (SpecCode,Title,ProductId,CreatedAt,CreatedBy) VALUES (?,?,?,?,?)',
                    (code, title, product_id, ts, "seed"))
        spec_id = cur.lastrowid
        # tạo version 1 ở trạng thái Draft
        cur.execute('INSERT INTO SpecVersions (SpecId,VersionNo,Status,CreatedAt,CreatedBy) VALUES (?,?,?,?,?)',
                    (spec_id, 1, "Draft", ts, "seed"))
        n_spec += 1

    for row in rows:
        g = lambda k: (str(row.get(k)).strip() if row.get(k) is not None else "")
        cust_id = get_or_create_customer(g("customer_code"), g("customer_name"))
        prod_id = get_or_create_product(g("product_code"), g("product_name"), cust_id)
        if g("spec_code"):
            ensure_spec(g("spec_code"), g("spec_title") or g("product_name"), prod_id)

    conn.commit()
    conn.close()
    print(f"✅ Nạp xong: +{n_cust} customer, +{n_prod} product, +{n_spec} spec (đã bỏ qua bản trùng).")


def main():
    ap = argparse.ArgumentParser(description="Nạp master data từ Excel/CSV vào SQLite MES.")
    ap.add_argument("path", help="File .xlsx hoặc .csv chứa master data.")
    ap.add_argument("--db", default="src/CCL.MES.Web/ccl_mes.db", help="Đường dẫn file SQLite.")
    args = ap.parse_args()
    upsert(args.path, args.db)


if __name__ == "__main__":
    main()
