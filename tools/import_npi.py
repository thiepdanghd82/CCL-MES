#!/usr/bin/env python3
"""
import_npi.py — Load NPI master data from the Data/ folder into the
CCL-MES SQLite database.

Loads 4 tables:
  - WorkCenters             (derived: distinct work centers from RoutingOperations)
  - RawMaterials            (from "Raw Materials.xlsx")
  - RoutingOperations       (from "RoutingOperations *.csv")
  - ManufacturingStructures (from "ManufacturingStructures *.csv")

Idempotency:
  Each table is cleared then refilled INSIDE a single transaction
  per run. Re-running produces the same end state (DELETE + INSERT
  pattern). On any exception, the transaction is rolled back — the
  DB stays exactly as it was before the run.

Per-table counters:
  imported = rows inserted into the table
  skipped  = rows present in the source but rejected at validation
             (missing required field, malformed row, header row,
             empty line, etc.)
  failed   = rows that triggered an exception during insert. With a
             single transactional batch insert, "failed" means the
             whole batch failed and was rolled back; the counter
             surfaces the row count that would have been written.

Prerequisites:
  1. Run `dotnet run --project src/CCL.MES.Web` once so EF Core creates
     the schema (including the 4 NPI tables) — file ccl_mes.db must
     contain `WorkCenters`, `RawMaterials`, `RoutingOperations`,
     `ManufacturingStructures` before this script runs.

Usage:
  pip install openpyxl --break-system-packages
  python3 tools/import_npi.py --data "<path-to-Data-folder>" --db src/CCL.MES.Web/ccl_mes.db
"""
import argparse
import csv
import glob
import os
import sqlite3
import sys
from datetime import datetime, timezone

csv.field_size_limit(10_000_000)

NPI_TABLES = ("WorkCenters", "RawMaterials", "RoutingOperations", "ManufacturingStructures")


def now():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")


def num(v):
    """Coerce a CSV/xlsx cell into a float; return 0.0 for blanks/dashes."""
    if v is None:
        return 0.0
    s_ = str(v).strip().replace(",", "")
    if s_ == "" or s_ == "-":
        return 0.0
    try:
        return float(s_)
    except ValueError:
        return 0.0


def s(v):
    """Coerce a cell into a trimmed string or None for empty."""
    if v is None:
        return None
    t = str(v).strip()
    return t if t != "" else None


def find(data_dir, pattern):
    hits = sorted(glob.glob(os.path.join(data_dir, pattern)))
    return hits[0] if hits else None


def infer_area(code, desc):
    t = f"{code} {desc}".upper()
    rules = [
        (("CNC",), "CNC"),
        (("AOI", "INSPECT", "AVT"), "INSPECTION"),
        (("SILK", "SS", "ARSS", "ASS", "MSS"), "SILKSCREEN"),
        (("FLEXO", "GALLUS", "BROTECH", "BFL", "GFL"), "FLEXO"),
        (("LASER",), "LASER"),
        (("PRESS", "PPSC"), "PRESS"),
        (("OVEN", "OVS", "BAKING"), "OVEN"),
        (("FB", "DIECUT", "CUT"), "DIECUT"),
        (("INDIGO", "IDG", "HP"), "DIGITAL"),
        (("MAN", "MANUAL", "PACK", "FQC", "OQC"), "MANUAL"),
        (("PRE", "FXPP"), "PRE-PRESS"),
        (("MAG", "LAM"), "LAMINATE"),
    ]
    for keys, area in rules:
        if any(k in t for k in keys):
            return area
    return "OTHER"


# ─── Reader helpers ──────────────────────────────────────────────
# Each reader returns the rows to insert + a counters dict carrying
# "seen", "skipped", and skip-reason histogram so callers can log
# the per-table breakdown BEFORE attempting the insert. "imported"
# + "failed" are determined at insert time inside the transaction.

def read_routing(path):
    """RoutingOperations + collect distinct work centers as a side-effect.

    Phase 7 hạng mục 2 — mở rộng từ 10 lên 20 cột để khớp CMES tham chiếu:
    4 numeric đổi sang num_or_none() (phân biệt 0 vs missing) + 10 field
    mới (Unit/Crew/SetupCrew/LaborClass/Alt/Effectivity/Efficiency/Site/
    RoutingType/Planner).

    Column map (IFS export "RoutingOperations *.csv", 62 columns):
        [0]  Create Date            (unused)
        [1]  Part No                → PartNo
        [2]  Operation No           → OpNo
        [3]  Part Description       → PartDescription
        [4]  Operation Description  → Operation
        [5]  Work Centre No         → WorkCenterNo
        [6]  Setup Labour Class     (unused)
        [7]  Labour Class           → LaborClass               (Phase 7 hm2)
        [8]  Work Centre Desc       → WorkCenterDescription
        [9]  Mach Setup Time        → MachineSetupTime         (Phase 7 hm2: → double?)
        [10] Labour Setup Time      → LaborSetupTime           (Phase 7 hm2: → double?)
        [11] Mach Run Factor        → MachineRunTime           (Phase 7 hm2: → double?)
        [12] Labour Run Factor      → LaborRunTime             (Phase 7 hm2: → double?)
        [13] Factor Unit            → Unit                     (Phase 7 hm2)
        [19] Setup Crew Size        → SetupCrew                (Phase 7 hm2)
        [20] Crew Size              → Crew                     (Phase 7 hm2)
        [21] Alternative            → Alt                      (Phase 7 hm2)
        [24] Routing Effectivity    → Effectivity              (Phase 7 hm2)
        [26] Planner                → Planner                  (Phase 7 hm2, parity Structure)
        [43] Efficiency Factor      → Efficiency               (Phase 7 hm2)
        [58] Site                   → Site                     (Phase 7 hm2)
        [60] Routing Type           → RoutingType              (Phase 7 hm2)
    """
    counters = {"seen": 0, "skipped": 0, "skip_reasons": {}}
    rows = []
    wc = {}
    if not path:
        return rows, wc, counters
    ts = now()
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.reader(f)
        next(rd, None)  # header row
        for r in rd:
            counters["seen"] += 1
            if len(r) < 13:
                counters["skipped"] += 1
                counters["skip_reasons"]["short_row"] = (
                    counters["skip_reasons"].get("short_row", 0) + 1
                )
                continue
            part = s(r[1])
            if not part:
                counters["skipped"] += 1
                counters["skip_reasons"]["missing_partno"] = (
                    counters["skip_reasons"].get("missing_partno", 0) + 1
                )
                continue
            wcno, wcdesc = s(r[5]), s(r[8])
            if wcno and wcno not in wc:
                wc[wcno] = wcdesc or ""
            # Safe-index helper — IFS exports may truncate trailing columns.
            def col(idx):
                return r[idx] if idx < len(r) else None
            rows.append(
                (
                    part,
                    s(r[3]),
                    s(r[2]),
                    s(r[4]),
                    wcno,
                    wcdesc,
                    num_or_none(r[9]),
                    num_or_none(r[10]),
                    num_or_none(r[11]),
                    num_or_none(r[12]),
                    # Phase 7 hạng mục 2 — 10 cột mới.
                    s(col(13)),    # Unit
                    num_or_none(col(20)),  # Crew (CSV col 21 → idx 20)
                    num_or_none(col(19)),  # SetupCrew (CSV col 20 → idx 19)
                    s(col(7)),     # LaborClass
                    s(col(21)),    # Alt
                    s(col(24)),    # Effectivity
                    num_or_none(col(43)),  # Efficiency
                    s(col(58)),    # Site
                    s(col(60)),    # RoutingType
                    s(col(26)),    # Planner
                    ts,
                )
            )
    return rows, wc, counters


def num_or_none(v):
    """Coerce a cell into float OR None for empty/non-numeric. Diff
    from num() which coerces empty/non-numeric → 0.0; used for the
    nullable double? columns added in Phase 7 hạng mục 1."""
    if v is None:
        return None
    t = str(v).strip().rstrip("%").replace(",", "")
    if t == "" or t == "-":
        return None
    try:
        return float(t)
    except ValueError:
        return None


def read_structures(path):
    """ManufacturingStructures (BOM).

    Column map (IFS export "ManufacturingStructures *.csv", 31 columns):
        [0]  Parent Part No              → ParentPart
        [1]  Parent Part Description     → ParentDescription
        [2]  Component Part              → ComponentPart
        [3]  Component Part Description  → ComponentDescription
        [4]  Parent Part Status          (unused)
        [5]  Qty Per Assembly            → QtyAssembly (numeric)
        [6]  Component Scrap             → ScrapFactor (numeric)
        [7]  Scrap Factor (%)            → ScrapPct  (Phase 7: double?)
        [8]  Pitch                       → Pitch     (Phase 7: double?)
        [9]  Cavity                      → Cavity    (Phase 7: double?)
        [10] Color Nums                  → Color
        [11] Phase In                    → EffectivityDate  (Phase 7)
        [14] Structure Effectivity       → Effectivity      (Phase 7)
        [15] Alternative No              → Alt              (Phase 7)
        [21] Structure Type              → StructureType    (Phase 7)
        [28] Planner                     → Planner          (Phase 7)
        [29] UOM                         → Uom

    Phase 7 hạng mục 1 — bổ sung 5 field mới (StructureType / Alt /
    Effectivity / EffectivityDate / Planner) + đổi 3 field Pitch /
    Cavity / ScrapPct từ string sang double? (Q9 verified 100% numeric
    trên 20,530 rows). Khớp 16 cột UI CMES tham chiếu.
    """
    counters = {"seen": 0, "skipped": 0, "skip_reasons": {}}
    rows = []
    if not path:
        return rows, counters
    ts = now()
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.reader(f)
        next(rd, None)
        for r in rd:
            counters["seen"] += 1
            if len(r) < 11:
                counters["skipped"] += 1
                counters["skip_reasons"]["short_row"] = (
                    counters["skip_reasons"].get("short_row", 0) + 1
                )
                continue
            parent = s(r[0])
            comp = s(r[2])
            if not parent and not comp:
                counters["skipped"] += 1
                counters["skip_reasons"]["no_parent_or_component"] = (
                    counters["skip_reasons"].get("no_parent_or_component", 0) + 1
                )
                continue
            uom = s(r[29]) if len(r) > 29 else None
            structure_type = s(r[21]) if len(r) > 21 else None
            alt = s(r[15]) if len(r) > 15 else None
            effectivity = s(r[14]) if len(r) > 14 else None
            effectivity_date = s(r[11]) if len(r) > 11 else None
            planner = s(r[28]) if len(r) > 28 else None
            rows.append(
                (
                    parent or "",
                    s(r[1]),
                    comp or "",
                    s(r[3]),
                    num(r[5]),
                    uom,
                    num(r[6]),
                    num_or_none(r[7]),
                    num_or_none(r[8]),
                    num_or_none(r[9]),
                    s(r[10]),
                    structure_type,
                    alt,
                    effectivity,
                    effectivity_date,
                    planner,
                    ts,
                )
            )
    return rows, counters


def read_raw_materials(path):
    """RawMaterials (IFS Raw Material catalog).

    Column map (IFS export "Raw Materials.xlsx", 69 columns):
        [0]  Part No                       → PartNo
        [1]  Part Description              → PartDescription
        [2]  Supplier ID                   → SupplierId
        [3]  Supplier Name                 → SupplierName
        [4]  Price                         → Price (numeric)
        [6]  Currency                      → Currency
        [7]  Price Unit Measure            → PriceUom
        [29] Status Code                   → Grp
        [30] Status Code Description       → Type / TypeDesc (mirrored)
        [31] Acquisition Type              → CatalogGroup
        [10] Site                          → CatalogDesc (numeric site code as string)

    The current IFS export does NOT carry the historical "Catalog Group"
    /"Catalog Desc" columns the entity schema originally targeted. The
    closest-meaning columns above are used so the read-side UI has
    something to show; if these mappings are wrong for business use,
    flag for Phase-1 follow-up and re-run the import (idempotent).
    """
    counters = {"seen": 0, "skipped": 0, "skip_reasons": {}}
    rows = []
    if not path:
        return rows, counters
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit(
            "ERROR: openpyxl is required. "
            "Install via 'pip install openpyxl --break-system-packages'."
        )
    ts = now()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header — not counted in "seen"
        counters["seen"] += 1
        if not r or r[0] is None:
            counters["skipped"] += 1
            counters["skip_reasons"]["missing_partno"] = (
                counters["skip_reasons"].get("missing_partno", 0) + 1
            )
            continue
        g = lambda k: r[k] if k < len(r) else None
        rows.append(
            (
                s(g(0)),
                s(g(1)),
                s(g(2)),
                s(g(3)),
                num(g(4)),
                s(g(6)),
                s(g(7)),
                s(g(31)),
                s(g(10)),
                s(g(29)),
                s(g(30)),
                s(g(30)),
                ts,
            )
        )
    return rows, counters


# ─── Insert helpers (called inside one transaction) ──────────────

def insert_workcenters(cur, wc_dict):
    ts = now()
    rows = [(code, desc, infer_area(code, desc), ts) for code, desc in sorted(wc_dict.items())]
    cur.executemany(
        'INSERT INTO "WorkCenters" ("Code","Description","Area","CreatedAt") VALUES (?,?,?,?)',
        rows,
    )
    return len(rows)


def insert_routing(cur, rows):
    # Phase 7 hạng mục 2 — 21 columns (20 data + CreatedAt). Order khớp
    # exactly với tuple shape trong read_routing.
    cur.executemany(
        'INSERT INTO "RoutingOperations" '
        '("PartNo","PartDescription","OpNo","Operation","WorkCenterNo","WorkCenterDescription",'
        '"MachineSetupTime","LaborSetupTime","MachineRunTime","LaborRunTime",'
        '"Unit","Crew","SetupCrew","LaborClass","Alt","Effectivity","Efficiency","Site","RoutingType","Planner",'
        '"CreatedAt") '
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def insert_structures(cur, rows):
    # Phase 7 hạng mục 1 — 17 columns (16 data + CreatedAt).
    cur.executemany(
        'INSERT INTO "ManufacturingStructures" '
        '("ParentPart","ParentDescription","ComponentPart","ComponentDescription",'
        '"QtyAssembly","Uom","ScrapFactor","ScrapPct","Pitch","Cavity","Color",'
        '"StructureType","Alt","Effectivity","EffectivityDate","Planner","CreatedAt") '
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def insert_raw_materials(cur, rows):
    cur.executemany(
        'INSERT INTO "RawMaterials" '
        '("PartNo","PartDescription","SupplierId","SupplierName","Price","Currency","PriceUom",'
        '"CatalogGroup","CatalogDesc","Grp","Type","TypeDesc","CreatedAt") '
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        rows,
    )
    return len(rows)


def table_count(cur, table):
    cur.execute(f'SELECT COUNT(*) FROM "{table}";')
    return cur.fetchone()[0]


def assert_schema(cur):
    """Bail early if the 4 NPI tables aren't in the schema."""
    cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
    have = {r[0] for r in cur.fetchall()}
    missing = [t for t in NPI_TABLES if t not in have]
    if missing:
        sys.exit(
            f"ERROR: NPI tables missing from schema: {missing}.\n"
            "       Run `dotnet run --project src/CCL.MES.Web` once so EF Core\n"
            "       creates the schema, then re-run this script."
        )


def format_skip_reasons(reasons):
    if not reasons:
        return "(none)"
    return ", ".join(f"{k}={v}" for k, v in sorted(reasons.items()))


def main():
    ap = argparse.ArgumentParser(description="Load NPI master data into the CCL-MES SQLite database.")
    ap.add_argument("--data", required=True, help="Path to the Data/ folder (source files).")
    ap.add_argument("--db", default="src/CCL.MES.Web/ccl_mes.db", help="Path to the SQLite file.")
    args = ap.parse_args()

    if not os.path.exists(args.db):
        sys.exit(
            f"ERROR: Database file not found at {args.db}.\n"
            "       Run `dotnet run --project src/CCL.MES.Web` once so EF Core\n"
            "       creates the schema, then re-run this script."
        )
    if not os.path.isdir(args.data):
        sys.exit(f"ERROR: Data folder not found: {args.data}")

    routing_csv = find(args.data, "RoutingOperations*.csv")
    struct_csv = find(args.data, "ManufacturingStructures*.csv")
    raw_xlsx = find(args.data, "Raw Materials*.xlsx") or find(args.data, "*aw*aterial*.xlsx")

    print(f"DB        : {args.db}")
    print(f"Data dir  : {args.data}")
    print(f"  routing : {routing_csv or '(not found — RoutingOperations will be empty)'}")
    print(f"  struct  : {struct_csv or '(not found — ManufacturingStructures will be empty)'}")
    print(f"  raw mat : {raw_xlsx or '(not found — RawMaterials will be empty)'}")
    print("")

    # Open + assert NPI tables exist.
    conn = sqlite3.connect(args.db)
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = OFF;")
    assert_schema(cur)

    # ── BEFORE counts (per NPI table) ─────────────────────────────
    before = {t: table_count(cur, t) for t in NPI_TABLES}
    print("BEFORE row counts:")
    for t in NPI_TABLES:
        print(f"  {t:<30} {before[t]:>10,}")
    print("")

    # ── READ sources (validation + skip-counting happen here) ────
    print("Reading sources...")
    routing_rows, wc_dict, c_rt = read_routing(routing_csv)
    print(
        f"  routing  : seen={c_rt['seen']:,}  skipped={c_rt['skipped']:,}  "
        f"reasons={format_skip_reasons(c_rt['skip_reasons'])}"
    )
    struct_rows, c_st = read_structures(struct_csv)
    print(
        f"  struct   : seen={c_st['seen']:,}  skipped={c_st['skipped']:,}  "
        f"reasons={format_skip_reasons(c_st['skip_reasons'])}"
    )
    raw_rows, c_rm = read_raw_materials(raw_xlsx)
    print(
        f"  raw mat  : seen={c_rm['seen']:,}  skipped={c_rm['skipped']:,}  "
        f"reasons={format_skip_reasons(c_rm['skip_reasons'])}"
    )
    print(f"  work ctr : derived={len(wc_dict):,} distinct codes from routing")
    print("")

    # ── WRITE inside a single transaction ────────────────────────
    print("Writing to DB (single transaction)...")
    imported = {t: 0 for t in NPI_TABLES}
    failed = {t: 0 for t in NPI_TABLES}

    try:
        cur.execute("BEGIN;")
        for t in NPI_TABLES:
            cur.execute(f'DELETE FROM "{t}";')
        imported["RoutingOperations"] = insert_routing(cur, routing_rows)
        imported["WorkCenters"] = insert_workcenters(cur, wc_dict)
        imported["ManufacturingStructures"] = insert_structures(cur, struct_rows)
        imported["RawMaterials"] = insert_raw_materials(cur, raw_rows)
        conn.commit()
    except Exception as exc:
        conn.rollback()
        # If any insert blew up, the batch that triggered it was lost.
        # Surface the row counts each batch would have written so the
        # operator sees the full impact in the final report.
        failed["RoutingOperations"] = len(routing_rows) if imported["RoutingOperations"] == 0 else 0
        failed["WorkCenters"] = len(wc_dict) if imported["WorkCenters"] == 0 else 0
        failed["ManufacturingStructures"] = (
            len(struct_rows) if imported["ManufacturingStructures"] == 0 else 0
        )
        failed["RawMaterials"] = len(raw_rows) if imported["RawMaterials"] == 0 else 0
        print(f"\nERROR: transaction rolled back — no rows persisted.\n  cause: {exc}\n")
        after = {t: table_count(cur, t) for t in NPI_TABLES}
        conn.close()
        _print_report(before, c_rt, c_st, c_rm, wc_dict, imported, failed, after, rolled_back=True)
        sys.exit(1)

    # ── AFTER counts ─────────────────────────────────────────────
    after = {t: table_count(cur, t) for t in NPI_TABLES}
    conn.close()

    _print_report(before, c_rt, c_st, c_rm, wc_dict, imported, failed, after, rolled_back=False)


def _print_report(before, c_rt, c_st, c_rm, wc_dict, imported, failed, after, *, rolled_back):
    print("")
    print("═══════════════════════════════════════════════════════════════════════")
    print("IMPORT REPORT")
    print("═══════════════════════════════════════════════════════════════════════")
    print(
        f"{'table':<28} {'before':>10} {'seen':>10} {'skipped':>10} "
        f"{'imported':>10} {'failed':>10} {'after':>10}"
    )
    rows = [
        ("RoutingOperations", c_rt["seen"], c_rt["skipped"]),
        ("WorkCenters", len(wc_dict), 0),
        ("ManufacturingStructures", c_st["seen"], c_st["skipped"]),
        ("RawMaterials", c_rm["seen"], c_rm["skipped"]),
    ]
    for t, seen, skipped in rows:
        print(
            f"{t:<28} {before[t]:>10,} {seen:>10,} {skipped:>10,} "
            f"{imported[t]:>10,} {failed[t]:>10,} {after[t]:>10,}"
        )
    print("═══════════════════════════════════════════════════════════════════════")
    if rolled_back:
        print("STATUS: ROLLED BACK — DB unchanged (the 'after' column above proves it).")
    else:
        print("STATUS: COMMITTED — all 4 NPI tables refilled atomically.")
    print("")


if __name__ == "__main__":
    main()
