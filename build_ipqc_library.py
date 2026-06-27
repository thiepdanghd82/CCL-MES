# -*- coding: utf-8 -*-
"""Build IPQC master library (import-ready) for CMES app.
One row per check-item. Coverage = full 86 items, with Short-form flag.
Source: IPQC_Checklist_TheoNhom_SongNgu_VN-EN_v1 + Pareto defect catalog 2025-2026.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- severity -> AQL map -------------------------------------------------
AQL = {"Critical": "0.65", "Major": "1.5", "Minor": "4.0"}
SYM = {"Critical": "◆", "Major": "●", "Minor": "○"}

# LIBRARY rows. Tuple:
# (Code, Group, Name_VI, Name_EN, Accept_VI, Accept_EN, Method, Sev,
#  InspType, DefectCode, Pareto, Short, ISO, Cond, Note)
G_A, G_B, G_C, G_D = "A·Ngoại quan", "B·Kích thước", "C·Màu sắc", "D·Chức năng"

LABEL = [
 ("A1",G_A,"Đúng nội dung in (text/logo/mã/seri)","Print content correct","Nội dung, chữ, logo, mã, seri ĐÚNG so spec/mẫu chuẩn","Content/code/serial correct vs master","Soi mắt + đối chiếu mẫu/spec; kính lúp","Critical","Visual","CONTENT","",True,"","","Lỗi lặp — sai phôi/trùng seri"),
 ("A2",G_A,"Lệch chồng màu (Mis-register)","Mis-register","Chồng màu khít so mẫu; lệch ≤ dung sai spec","Registration tight; ≤ tol","Soi mắt + đối chiếu mẫu; kính lúp","Critical","Visual","MISREG","40.3",True,"","","Top1 — 40% NG"),
 ("A3",G_A,"Bẩn / đốm / chấm (Dirty/Smear/Dot)","Dirty/Smear/Dot","Không vết bẩn, đốm mực, dây mực, chấm lạ","No dirt/smear/stray dots","Soi mắt dưới đèn chuẩn","Critical","Visual","DIRTY","27.3",True,"","","Top2 — Dirty 26% + Smear 1.3%"),
 ("A4",G_A,"Loang, sọc (Mottled/Streak)","Mottled/Streak","Lớp mực đều, không loang, sọc","Even ink, no mottle/streak","Soi mắt nghiêng + đèn","Major","Visual","MOTTLE","2.1",True,"","",""),
 ("A5",G_A,"Xước bề mặt (Scratch)","Scratch","Không vết xước trên bề mặt in","No scratch on print","Soi mắt nghiêng + đèn","Major","Visual","SCRATCH","6.0",True,"","",""),
 ("A6",G_A,"Bavia, dính/tràn keo (Burr/Glue)","Burr/Glue","Mép sạch, không bavia, không dính/tràn keo","Clean edge, no burr/glue","Soi mắt + sờ mép","Major","Visual","BURR","10.1",True,"","",""),
 ("A7",G_A,"Bong / phồng ép dán (Peel/Blister)","Peel/Blister","Lớp cán/ép dán bám chắc, không bong, phồng","Laminate bonded, no peel","Soi mắt + gập nhẹ","Major","Visual","PEEL","1.9",False,"","Có ép dán/cán","If laminated"),
 ("A8",G_A,"Cắt sâu / thủng lót (Full-cut)","Full-cut/Liner puncture","Bế đúng lớp, không thủng đế, không cắt-xẻ lệch","Die-cut correct layer, no liner puncture","Soi mắt + tách thử","Major","Visual","FULLCUT","4.2",True,"","",""),
 ("A9",G_A,"Lõm, gãy, hằn (Dent/Crease)","Dent/Crease","Không lõm, gãy, hằn trên sản phẩm","No dent/crack/crease","Soi mắt","Major","Visual","DENT","4.1",True,"","",""),
 ("A10",G_A,"Thiếu áp lực / mất nét","Low pressure/Missing detail","In đủ nét, rõ ràng, không mất chi tiết","Sharp print, no missing detail","Soi mắt + kính lúp; vs mẫu","Critical","Visual","LOWPRESS","1.0",True,"","","Lỗi lọt — khuôn bẩn"),
 ("A11",G_A,"Không đều (Uneven coverage)","Uneven coverage","Lớp in/phủ đều, không chỗ mỏng/dày","Even coverage","Soi mắt nghiêng","Major","Visual","UNEVEN","0.8",False,"","",""),
 ("A12",G_A,"Khuyết, thiếu chi tiết (Incomplete)","Incomplete","Đủ chi tiết theo mẫu chuẩn","All details present","Soi mắt + đối chiếu","Major","Visual","MISSING","1.0",False,"","",""),
 ("A13",G_A,"Nhăn / biến dạng (Wrinkle/Deform)","Wrinkle/Deform","Bề mặt phẳng, không nhăn, biến dạng","Flat, no wrinkle/deform","Soi mắt","Major","Visual","OTHER","",False,"","","Lỗi lọt FQC"),
 ("A14",G_A,"Lỗ pin (Pin hole)","Pin hole","Không lỗ kim trên lớp in","No pinhole","Soi mắt + đèn nền","Minor","Visual","PINHOLE","0.1",False,"","",""),
 ("A15",G_A,"Hoàn thiện bề mặt (gloss/matte/laminate)","Finish","Lớp phủ/cán/vân đúng loại & đều theo mẫu","Finish correct & even","Soi mắt + sờ; vs mẫu","Minor","Visual","OTHER","",False,"","Có phủ/cán","If coated"),
 ("B1",G_B,"Kích thước tổng thể (W×L)","Overall size","Trong dung sai bản vẽ","Within drawing tol","Thước cặp/CMM (KHÔNG thước lá cho dung sai nhỏ)","Critical","Measure","SIZE","0.6",True,"","","Lỗi lọt — sai dụng cụ đo"),
 ("B2",G_B,"Kích thước vùng in (Print area)","Print area","Đúng kích thước & tỉ lệ spec","Correct size & ratio","Thước + kính lúp đo","Major","Measure","SIZE","",False,"","",""),
 ("B3",G_B,"Vị trí in / lề / căn giữa","Position/Margin","Đúng vị trí so đế & đường bế; lề đều","Correct position; even margins","Thước + đối chiếu layout","Major","Measure","SIZE","",False,"","",""),
 ("B4",G_B,"Bước nhảy / khoảng lặp (Step/Repeat)","Step/Repeat","Đúng bước, không sai khoảng lặp","Correct step/repeat","Thước + đếm; TẮT chế độ mắt đọc khi kiểm","Critical","Measure","SIZE","",True,"","","Lỗi lặp — sai bước nhảy"),
 ("B5",G_B,"Kích thước & hình bế, góc dao","Die shape/size/angle","Đúng hình bế, kích thước & góc dao spec","Correct die shape/size/angle","Đối chiếu mẫu/dưỡng; cắt thử dao","Critical","Measure","SIZE","",True,"","","Lỗi lặp — sai góc dao"),
 ("B6",G_B,"Khoảng cách giữa con / lỗ định vị","Gap/Hole","Đúng khoảng cách giữa con & vị trí lỗ","Correct gap & hole pos","Thước + đối chiếu dưỡng","Major","Measure","SIZE","",False,"","",""),
 ("B7",G_B,"Số con/hàng trên cuộn-tờ (Count)","Count","Đúng số con/hàng theo spec","Correct count","Đếm + đối chiếu","Minor","Measure","OTHER","",False,"","",""),
 ("C1",G_C,"Màu so mẫu chuẩn (ΔE)","Color vs master","Đạt trong giới hạn màu; ΔE ≤ spec (mặc định ≤2)","Within color limits; ΔE ≤ spec","Đèn chuẩn D50/D65 + máy đo màu (nếu có)","Critical","Visual","COLOR","0.1",True,"","",""),
 ("C2",G_C,"Mã mực đúng spec (Ink code)","Ink code","Mã mực thực tế ↔ spec khớp","Ink code matches spec","Đối chiếu tem mực ↔ spec","Major","Visual","OTHER","",False,"","","Lỗi lặp (In lụa)"),
 ("C3",G_C,"Đồng đều màu trong & giữa con","Uniformity","Màu đều toàn tem & giữa các con","Color even","Soi mắt nhiều vị trí","Major","Visual","OTHER","",False,"","",""),
 ("C4",G_C,"Độ đậm / độ phủ mực (Density)","Density/Coverage","Đậm nhạt & độ phủ trong giới hạn mẫu","Density within limits","Soi mắt; densitometer (nếu có)","Major","Visual","OTHER","",False,"","",""),
 ("C5",G_C,"Độ bóng / mờ bề mặt (Gloss/Matte)","Gloss/Matte","Độ bóng/mờ đúng loại spec, đồng nhất","Gloss/matte per spec","Soi mắt nghiêng; vs mẫu","Major","Visual","OTHER","",False,"","",""),
 ("C6",G_C,"Đồng nhất màu giữa các lô","Lot-to-lot","Màu nhất quán so lô trước & mẫu","Consistent vs prev lot","Đối chiếu mẫu lưu lô trước","Minor","Visual","OTHER","",False,"","",""),
 ("D1",G_D,"Độ bám mực — Tape test","Tape adhesion","Dán & bóc băng keo 3M: mực không bong","Apply/peel tape: no ink loss","Băng keo chuẩn, bóc 1 lần dứt khoát","Critical","Functional","OTHER","",True,"","","Phá hủy; theo spec khách"),
 ("D2",G_D,"Cross-cut — rạch ô bàn cờ","Cross-cut","ISO 2409 ≤ Class 1 / ASTM D3359 ≥ 4B","ISO2409 ≤1 / D3359 ≥4B","Dao rạch ô + băng keo, bóc & soi","Critical","Functional","OTHER","",True,"ISO 2409/ASTM D3359","",""),
 ("D3",G_D,"Mài mòn cồn (Alcohol/IPA rub)","IPA rub","Chà cồn theo spec (vd 50 lần/500g): không mờ, mất","IPA rub per spec: no fade","Máy/bút chà + bông tẩm cồn","Critical","Functional","OTHER","",True,"","","Phá hủy; theo spec khách"),
 ("D4",G_D,"Mài mòn / chà xước (Rub/RCA)","Rub/RCA","Đạt số chu kỳ spec, không mòn lớp in","Pass cycles, no wear","RCA tester / eraser test","Major","Functional","OTHER","",False,"","",""),
 ("D5",G_D,"Đọc mã vạch (Barcode/QR)","Barcode/QR","Quét đạt cấp ≥ spec; đúng nội dung & seri, không trùng","Scan grade ≥ spec; no dup","Máy quét/verifier","Critical","Functional","BARCODE","0.0",True,"ISO 15415/15416","Có mã vạch","Lỗi lặp — trùng seri"),
 ("D6",G_D,"Độ bám keo dán (Adhesive/Peel)","Adhesive peel","Dán lên bề mặt chuẩn: không bong mép, đủ lực bóc","No edge lift, adequate peel","Dán thử + lực kế (nếu spec)","Major","Functional","OTHER","",False,"","Nhãn có keo","If adhesive"),
]

SILK = [
 ("A1",G_A,"Đúng nội dung in (text/logo/mã)","Print content correct","Nội dung, ký tự, logo, mã ĐÚNG so spec/mẫu","Content correct vs master","Soi mắt + đối chiếu mẫu/spec; kính lúp","Critical","Visual","CONTENT","",True,"","","Nội dung biến đổi"),
 ("A2",G_A,"Bẩn / bụi / đốm / chấm","Dirty/Dust/Smear/Dot","Không vết bẩn, bụi, đốm, chấm mực","No dirt/dust/smear/dots","Soi mắt dưới đèn chuẩn","Critical","Visual","DIRTY","72.6",True,"","","Top1 — 73% NG (áp đảo)"),
 ("A3",G_A,"Mất nét ký tự / Incomplete","Missing strokes","In đủ nét, đủ lớp, không mất nét chữ","Full sharp print, all layers","Soi mắt + kính lúp; vs mẫu","Critical","Visual","MISSING","1.2",True,"","","Lỗi lặp — khuôn/NL bẩn"),
 ("A4",G_A,"Lệch (Mis-register)","Mis-register","Chồng lớp/màu khít so mẫu","Layers registered","Soi mắt + đối chiếu","Major","Visual","MISREG","8.7",True,"","",""),
 ("A5",G_A,"Không đều / mỏng-dày (Uneven)","Uneven","Lớp mực đều, không chỗ mỏng/dày","Even ink","Soi mắt nghiêng","Major","Visual","UNEVEN","1.2",False,"","",""),
 ("A6",G_A,"Xước (Scratch)","Scratch","Không vết xước bề mặt in","No scratch","Soi mắt + đèn","Major","Visual","SCRATCH","1.4",True,"","",""),
 ("A7",G_A,"Bong mực (Peel-off ink)","Peel-off ink","Mực bám chắc, không bong","Ink adhered","Soi mắt + sờ","Major","Visual","PEEL","0.3",False,"","",""),
 ("A8",G_A,"Lỗ pin / chấm do thủng khuôn","Pinhole","Không chấm, lỗ do thủng khuôn","No pinhole from screen","Soi mắt + đèn nền","Major","Visual","PINHOLE","0.2",False,"","","Lỗi lặp — thủng khuôn"),
 ("A9",G_A,"Hiện hình / vệt (Ghost/Streak)","Ghost/Streak","Không vệt, bóng hình lạ trên SP","No ghost/stray image","Soi mắt","Major","Visual","OTHER","",False,"","","Lỗi lặp — lớp matte"),
 ("A10",G_A,"Lõm (Dent)","Dent","Không lõm trên bề mặt","No dent","Soi mắt nghiêng","Major","Visual","DENT","0.3",False,"","",""),
 ("A11",G_A,"Bavia / cắt sâu mép (Burr/Full-cut)","Burr/Full-cut","Mép sạch, không bavia, không cắt sâu","Clean edge, no over-cut","Soi mắt + sờ mép","Major","Visual","BURR","0.0",False,"","",""),
 ("B1",G_B,"Kích thước & vị trí in","Size/Position","Đúng vị trí, trong dung sai spec","Correct position & tol","Thước cặp + đối chiếu layout","Major","Measure","SIZE","",True,"","",""),
 ("B2",G_B,"Độ dày / độ phủ lớp in (Film/Opacity)","Film/Coverage","Đủ độ dày/độ chắn sáng (mực đèn) theo spec","Sufficient film/opacity","Soi sáng xuyên / đo (nếu spec)","Major","Measure","OTHER","",False,"","Mực đèn","Lỗi lặp — mực đèn lọt sáng"),
 ("B3",G_B,"Vị trí ký tự / khoảng cách","Char position","Ký tự đúng vị trí & khoảng cách spec","Char position & spacing","Thước + kính lúp; đối chiếu","Major","Measure","SIZE","",False,"","",""),
 ("B4",G_B,"Số lớp in / thứ tự lớp","Layers/Sequence","Đủ số lớp & đúng thứ tự in spec","All layers, correct sequence","Đối chiếu spec + đếm lớp","Major","Measure","OTHER","",False,"","",""),
 ("C1",G_C,"Màu so mẫu chuẩn (ΔE)","Color vs master","Đạt giới hạn màu; ΔE ≤ spec","Within color limits","Đèn chuẩn + máy đo màu (nếu có)","Critical","Visual","COLOR","0.4",True,"","",""),
 ("C2",G_C,"Mã mực đúng spec (Ink code)","Ink code","Mã mực thực tế ↔ spec khớp","Ink code matches spec","Đối chiếu tem mực ↔ spec","Critical","Visual","OTHER","",True,"","","Lỗi lặp — mã mực sai"),
 ("C3",G_C,"Độ sáng mực đèn / huỳnh quang","Luminous","Mực đèn đủ sáng, đều, không lọt sáng","Luminous bright, even","Soi buồng tối / đèn UV","Major","Visual","OTHER","",True,"","Mực đèn","Lỗi lặp — mực đèn"),
 ("C4",G_C,"Đồng đều màu (Uniformity)","Uniformity","Màu đều toàn SP & giữa các con","Color even","Soi mắt nhiều vị trí","Major","Visual","OTHER","",False,"","",""),
 ("C5",G_C,"Độ bóng / mờ (Gloss/Matte)","Gloss/Matte","Độ bóng/mờ đúng loại spec, đồng nhất","Gloss/matte per spec","Soi mắt nghiêng; vs mẫu","Major","Visual","OTHER","",False,"","",""),
 ("D1",G_D,"Độ bám mực — Tape test","Tape adhesion","Dán & bóc băng keo: mực không bong","Apply/peel tape: no loss","Băng keo chuẩn, bóc 1 lần","Critical","Functional","OTHER","",True,"","","Phá hủy; theo spec khách"),
 ("D2",G_D,"Cross-cut — rạch ô bàn cờ","Cross-cut","ISO 2409 ≤ Class 1 / ASTM D3359 ≥ 4B","ISO2409 ≤1 / D3359 ≥4B","Dao rạch ô + băng keo","Critical","Functional","OTHER","",True,"ISO 2409/ASTM D3359","",""),
 ("D3",G_D,"Mài mòn cồn (Alcohol/IPA rub)","IPA rub","Chà cồn theo spec (vd 50–100 lần): ký tự không mờ","IPA rub per spec: no fade","Máy/bút chà + bông tẩm cồn","Critical","Functional","OTHER","",True,"","","Quan trọng cho in lụa ký tự"),
 ("D4",G_D,"Mài mòn / chà xước (Rub/RCA)","Rub/RCA","Đạt số chu kỳ spec, không mòn lớp in","Pass cycles, no wear","RCA / eraser test","Major","Functional","OTHER","",False,"","",""),
 ("D5",G_D,"Đọc mã vạch (Barcode, nếu có)","Barcode","Quét đạt cấp ≥ spec, đúng nội dung","Scan grade ≥ spec","Máy quét/verifier","Major","Functional","BARCODE","",False,"ISO 15415/15416","Có mã vạch",""),
]

PNC = [
 ("A1",G_A,"Gãy / nứt (Crack)","Crack","Không gãy, nứt trên sản phẩm","No crack/fracture","Soi mắt + uốn/soi nhẹ","Critical","Visual","CRACK","16.3",True,"","","Top1 — 16% NG"),
 ("A2",G_A,"Sứt mẻ (Dent/Chip)","Dent/Chip","Không sứt, mẻ cạnh/bề mặt","No chip/nick","Soi mắt + sờ cạnh","Critical","Visual","DENT","12.7",True,"","",""),
 ("A3",G_A,"Lồi lõm (Convex/Concave)","Convex/Concave","Bề mặt phẳng đều, không lồi lõm","Flat even surface","Soi mắt nghiêng + thước thẳng","Major","Visual","CONVEX","10.3",True,"","",""),
 ("A4",G_A,"Bavia (Burr)","Burr","Cạnh cắt sạch, không bavia","Clean cut edge, no burr","Soi mắt + sờ cạnh","Major","Visual","BURR","10.8",True,"","",""),
 ("A5",G_A,"Sơ, dắt mép (Fray/Whisker)","Fray/Whisker","Không sơ, dắt mép cắt","No fray at cut edge","Soi mắt + kính lúp","Major","Visual","FRAY","9.8",True,"","",""),
 ("A6",G_A,"Vỡ sơn / nứt sơn (Paint crack)","Paint crack","Lớp sơn/in không vỡ, nứt","Paint/print no crack","Soi mắt + uốn nhẹ","Critical","Visual","PAINTCRK","7.3",True,"","","Lỗi nhầm hàng (2 mã giống design)"),
 ("A7",G_A,"Trầy xước (Scratch)","Scratch","Không vết trầy trên bề mặt","No scratch","Soi mắt + đèn","Major","Visual","SCRATCH","5.3",True,"","",""),
 ("A8",G_A,"Dập lệch / sai vị trí dập","Press deviation","Vị trí dập đúng, không lệch","Correct press position","Soi mắt + đối chiếu dưỡng","Major","Visual","PRESSDEV","2.6",False,"","",""),
 ("A9",G_A,"Loang / bẩn bề mặt (Dirty)","Dirty","Bề mặt sạch, không loang, bẩn","Clean, no stain","Soi mắt dưới đèn","Major","Visual","DIRTY","0.1",False,"","",""),
 ("A10",G_A,"Dính dầu (Oil stain)","Oil stain","Bề mặt sạch, không dính dầu","Clean, no oil","Soi mắt + lau thử","Major","Visual","OIL","1.1",False,"","",""),
 ("A11",G_A,"Đúng nội dung khắc/in","Engraved/printed content","Nội dung khắc/in ĐÚNG so spec/mẫu","Content correct (if any)","Soi mắt + đối chiếu mẫu/spec","Major","Visual","OTHER","",False,"","Có in/khắc","If printed/engraved"),
 ("A12",G_A,"NG keo / tràn keo (Glue NG)","Glue NG","Keo đúng vị trí/lượng, không tràn","Glue correct, no overflow","Soi mắt","Minor","Visual","GLUE","2.2",False,"","Có keo","If glued"),
 ("B1",G_B,"Kích thước tổng thể (Overall dim.)","Overall dim","Trong dung sai bản vẽ; đo TRƯỚC khi gấp/bóng kính","Within tol; measure BEFORE fold","Thước cặp / CMM","Critical","Measure","SIZE","",True,"","","Lỗi lặp — dao biến dạng"),
 ("B2",G_B,"Vị trí dập / cắt (Press/Cut pos.)","Press/Cut position","Đúng vị trí so dưỡng/bản vẽ","Correct vs gauge/drawing","Đối chiếu dưỡng + thước","Critical","Measure","SIZE","",True,"","",""),
 ("B3",G_B,"Độ sâu cắt / lớp lót (Cut depth)","Cut depth/Liner","Không cắt sâu vào lõi, không rách lớp lót","No over-cut into core","Tách & soi lớp lót; đo","Critical","Measure","FULLCUT","7.9",True,"","","Lỗi lặp — lõi cắt sâu"),
 ("B4",G_B,"Góc / biên dạng (Angle/Profile)","Angle/Profile","Đúng hình dao/khuôn, đúng góc","Correct die shape & angle","Đối chiếu dưỡng/mẫu","Major","Measure","SIZE","",False,"","",""),
 ("B5",G_B,"Vị trí / đường kính lỗ (Hole)","Hole pos./dia.","Lỗ đúng vị trí & đường kính spec","Holes correct pos & dia","Thước cặp / pin gauge","Major","Measure","SIZE","",False,"","Có lỗ","If holes"),
 ("B6",G_B,"Độ phẳng / cong vênh (Flatness/Warp)","Flatness/Warp","Phẳng, không cong vênh quá dung sai","Flat, warp within tol","Thước thẳng / feeler trên mặt phẳng","Major","Measure","SIZE","",False,"","",""),
 ("C1",G_C,"Màu sơn/in so mẫu (Color vs master)","Color vs master","Đạt giới hạn màu so mẫu chuẩn","Within color limits","Đèn chuẩn + máy đo (nếu có)","Major","Visual","OTHER","",False,"","Có sơn/in","If painted/printed"),
 ("C2",G_C,"Đồng đều bề mặt sơn (Uniformity)","Paint uniformity","Lớp sơn đều, không loang/khác sắc","Even paint","Soi mắt","Major","Visual","OTHER","",False,"","Có sơn","If painted"),
 ("C3",G_C,"Độ bóng / mờ lớp phủ (Gloss/Matte)","Gloss/Matte","Độ bóng/mờ đúng loại spec, đồng nhất","Gloss/matte per spec","Soi mắt nghiêng; vs mẫu","Major","Visual","OTHER","",False,"","Có phủ","If coated"),
 ("C4",G_C,"Đúng màu nền / mạ (Plating/base)","Plating/base color","Màu nền/mạ kim loại đúng spec, đều","Base/plating color per spec","Soi mắt + đối chiếu mẫu","Minor","Visual","OTHER","",False,"","Chi tiết kim loại","Metal part"),
 ("D1",G_D,"Độ bám sơn/mực — Cross-cut","Cross-cut adhesion","ISO 2409 ≤ Class 1 / ASTM D3359 ≥ 4B","ISO2409 ≤1 / D3359 ≥4B","Dao rạch ô + băng keo","Critical","Functional","PAINTCRK","",True,"ISO 2409/ASTM D3359","Có sơn/in","If painted/printed"),
 ("D2",G_D,"Mài mòn cồn (Alcohol/IPA rub)","IPA rub","Chà cồn theo spec: lớp sơn/in không mờ, mất","IPA rub: no fade/loss","Máy/bút chà + bông tẩm cồn","Critical","Functional","OTHER","",True,"","Có sơn/in",""),
 ("D3",G_D,"Lắp ghép / đóng mở (Fit/Assembly)","Fit/Assembly","Lắp khít, đóng mở/đúng chức năng spec","Fits/functions per spec","Lắp thử với mẫu đối ứng","Critical","Functional","OTHER","",True,"","","Chức năng cơ khí"),
 ("D4",G_D,"Độ bám / lực bóc keo (Peel)","Peel force","Bám chắc, đủ lực bóc theo spec","Adhered, adequate peel","Dán thử + lực kế (nếu spec)","Major","Functional","OTHER","",False,"","Có keo","If adhesive"),
 ("D5",G_D,"Chức năng cơ khí khác (Bend/Snap…)","Other mechanical","Đạt theo yêu cầu spec (uốn, gập, snap…)","Pass spec requirement","Test theo spec sản phẩm","Major","Functional","OTHER","",False,"","",""),
]

DIGITAL = [
 ("A1",G_A,"Đúng nội dung in & seri biến đổi","Print/variable content correct","Nội dung, mã, seri biến đổi ĐÚNG so spec/mẫu; không trùng seri","Content/variable serial correct, no dup","Soi mắt + đối chiếu mẫu/spec; kính lúp","Critical","Visual","CONTENT","",True,"","","Indigo in dữ liệu biến đổi"),
 ("A2",G_A,"Sọc ngang / banding","Banding","Tông đều, không sọc ngang theo chu kỳ blanket","Even tone, no periodic banding","Soi mắt nghiêng + đèn; đối chiếu mẫu","Major","Visual","BANDING","",True,"","","Lỗi đặc thù digital"),
 ("A3",G_A,"Sương nền / chấm spray / bẩn","Background fog / spray dots","Nền sạch, không sương mực, chấm spray, bẩn","Clean ground, no fog/spray/dirt","Soi mắt dưới đèn chuẩn","Major","Visual","DIRTY","",True,"","",""),
 ("A4",G_A,"Mất nét / vạch trắng / dropout","Dropout / white line","Đủ nét, không mất hạt mực, vạch trắng do blanket/ITM","Full coverage, no dropout/white line","Soi mắt + kính lúp; đối chiếu mẫu","Critical","Visual","DROPOUT","",True,"","","Blanket/ITM lỗi → mất in"),
 ("A5",G_A,"Bóng ma / lưu hình blanket (Ghosting)","Ghosting / blanket memory","Không bóng hình lặp từ frame trước","No repeated ghost image","Soi mắt","Major","Visual","GHOST","",True,"","","Đặc thù digital"),
 ("A6",G_A,"Lệch chồng màu (digital)","Mis-register","Các màu/lớp chồng khít so mẫu","Colors registered vs master","Soi mắt + kính lúp","Major","Visual","MISREG","",False,"","",""),
 ("A7",G_A,"Xước bề mặt (Scratch)","Scratch","Không vết xước trên bề mặt in","No scratch on print","Soi mắt nghiêng + đèn","Major","Visual","SCRATCH","",False,"","",""),
 ("B1",G_B,"Vị trí in / đặt ảnh (Image placement)","Image placement","Đúng vị trí & lề so spec/đường bế","Correct position & margins","Thước + đối chiếu layout","Major","Measure","SIZE","",False,"","",""),
 ("B2",G_B,"Chiều dài lặp / co giãn ảnh","Repeat length / scaling","Chiều dài ảnh/bước lặp trong dung sai (co giãn vật liệu)","Repeat length within tol","Thước + đếm bước; đối chiếu","Major","Measure","SIZE","",False,"","","Digital dễ trôi chiều dài"),
 ("C1",G_C,"Màu so mẫu chuẩn (ΔE)","Color vs master","Đạt giới hạn màu; ΔE ≤ spec (mặc định ≤2)","Within color limits; ΔE ≤ spec","Đèn chuẩn D50/D65 + máy đo màu","Critical","Visual","COLOR","",True,"","",""),
 ("C2",G_C,"Đồng nhất màu trong & giữa lô (drift)","Color consistency / drift","Màu ổn định trong job & giữa lô; không trôi màu","Stable color within/between runs","So mẫu lưu + đo đầu/giữa/cuối job","Major","Visual","COLORDRIFT","",True,"","","Digital trôi màu nhanh hơn"),
 ("D1",G_D,"Bám ElectroInk — Tape test","ElectroInk adhesion","Dán & bóc băng keo: mực không bong","Apply/peel tape: no ink loss","Băng keo chuẩn, bóc 1 lần dứt khoát","Critical","Functional","OTHER","",True,"","","Indigo cần corona/primer; phá hủy"),
 ("D2",G_D,"Cross-cut — rạch ô bàn cờ","Cross-cut","ISO 2409 ≤ Class 1 / ASTM D3359 ≥ 4B","ISO2409 ≤1 / D3359 ≥4B","Dao rạch ô + băng keo, bóc & soi","Critical","Functional","OTHER","",True,"ISO 2409/ASTM D3359","",""),
 ("D3",G_D,"Mài mòn cồn (Alcohol/IPA rub)","IPA rub","Chà cồn theo spec: chữ/mực không mờ, mất","IPA rub per spec: no fade","Máy/bút chà + bông tẩm cồn","Critical","Functional","OTHER","",True,"","","Phá hủy; theo spec khách"),
 ("D4",G_D,"Đọc mã vạch + seri biến đổi","Barcode/QR + variable serial","Quét đạt cấp ≥ spec; đúng nội dung & seri; không trùng","Scan grade ≥ spec; no dup serial","Máy quét/verifier 100% đầu chuyền","Critical","Functional","BARCODE","",True,"ISO 15415/15416","Có mã vạch/seri","Indigo dữ liệu biến đổi"),
]

LINES = [("LBL","IN NHÃN / LABEL",LABEL),("DGT","IN SỐ / DIGITAL",DIGITAL),("SLK","IN LỤA / SILK",SILK),("PNC","DẬP-CNC / PRESS-CNC",PNC)]

# ---- defect catalog (Pareto) -> NG reason dropdown -----------------------
DEFECTS = {
 "LBL":[("MISREG","Lệch (Mis-register)",34887,40.3),("DIRTY","Bẩn (Dirty)",22548,26.0),("BURR","Bavia, keo (Burr/Glue)",8785,10.1),("SCRATCH","Xước (Scratch)",5175,6.0),("FULLCUT","Cắt sâu, thủng (Full-cut)",3593,4.2),("DENT","Lõm/Gãy/Hằn (Dent)",3536,4.1),("MOTTLE","Loang, sọc (Mottled)",1802,2.1),("PEEL","Bong/phồng ép dán (Peel-off)",1647,1.9),("SMEAR","Đốm bẩn (Smear)",1094,1.3),("MISSING","Khuyết, thiếu (Missing)",893,1.0),("LOWPRESS","Thiếu áp lực/mất nét",883,1.0),("UNEVEN","Không đều (Uneven)",687,0.8),("SIZE","Kích thước (Size)",531,0.6),("OTHER","Lỗi khác (Other)",287,0.3),("COLOR","NG màu (Color)",120,0.1),("PINHOLE","Lỗ pin (Pin hole)",73,0.1),("BARCODE","Mã vạch (Barcode)",23,0.0),("CONTENT","Sai nội dung/seri (Content)",None,None)],
 "SLK":[("DIRTY","Bẩn (Dirty)",10488,72.6),("OTHER","Lỗi khác (Other)",2047,14.2),("MISREG","Lệch (Mis-register)",1252,8.7),("SCRATCH","Xước (Scratch)",209,1.4),("UNEVEN","Không đều (Uneven)",170,1.2),("COLOR","NG màu (Color)",57,0.4),("SMEAR","Đốm bẩn (Smear)",44,0.3),("PEEL","Bong mực (Peel-off)",44,0.3),("DENT","Lõm (Dent)",37,0.3),("PINHOLE","Lỗ pin (Pin hole)",28,0.2),("DUST","Bụi (Dust)",22,0.2),("MISSING","Mất nét/Incomplete",18,0.1),("FULLCUT","Cắt sâu (Full-cut)",8,0.1),("BARCODE","Mã vạch (Barcode)",7,0.0),("DOT","Chấm (Dot)",5,0.0),("BURR","Bavia (Burr)",2,0.0),("CONTENT","Sai nội dung (Content)",None,None)],
 "DGT":[("CONTENT","Sai nội dung/seri (Content)",None,None),("DROPOUT","Mất nét/vạch trắng (Dropout)",None,None),("BANDING","Sọc ngang (Banding)",None,None),("COLORDRIFT","Trôi màu / ΔE (Color drift)",None,None),("GHOST","Bóng ma (Ghosting)",None,None),("DIRTY","Sương nền/spray (Fog)",None,None),("MISREG","Lệch (Mis-register)",None,None),("COLOR","NG màu (Color)",None,None),("SCRATCH","Xước (Scratch)",None,None),("BARCODE","Mã vạch/seri (Barcode)",None,None),("OTHER","Lỗi khác (Other)",None,None)],
 "PNC":[("CRACK","Gãy (Crack)",853,16.3),("OTHER","Lỗi khác (Other)",716,13.7),("DENT","Sứt mẻ (Dent/Chip)",663,12.7),("BURR","Bavia (Burr)",563,10.8),("CONVEX","Lồi lõm (Convex)",541,10.3),("FRAY","Sơ, dắt (Fray)",510,9.8),("FULLCUT","Cắt sâu (Full-cut)",412,7.9),("PAINTCRK","Vỡ sơn (Paint crack)",382,7.3),("SCRATCH","Trầy xước (Scratch)",275,5.3),("PRESSDEV","Dập lệch (Press deviated)",136,2.6),("GLUE","NG keo (Glue)",114,2.2),("OIL","Dính dầu (Oil)",57,1.1),("DIRTY","Loang (Dirty)",7,0.1),("SIZE","Kích thước (Size)",None,None)],
}

# ---- styling helpers -----------------------------------------------------
NAVY = "1F3864"; BLUE="2E5496"; LBLUE="D9E1F2"; LGREY="F2F2F2"
CR="C00000"; MA="BF8F00"; MI="808080"
VITAL="FCE4D6"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin,right=thin,top=thin,bottom=thin)
H = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HC = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
B = Font(name="Calibri", bold=True, size=10)
N = Font(name="Calibri", size=9)
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

# ===================== GUIDE =====================
ws = wb.active; ws.title="GUIDE"
ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=2
ws.column_dimensions["B"].width=120
def gline(r,txt,font=None,fill=None):
    c=ws.cell(row=r,column=2,value=txt)
    if font: c.font=font
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(wrap_text=True,vertical="center")
    return c
ws.row_dimensions[2].height=28
gline(2,"THƯ VIỆN IPQC — CHECK-ITEM CHO APP CMES (Song ngữ VN/EN)",HC,NAVY)
gline(3,"IPQC Inspection Library by Product Line · CCL Vietnam — Yen Phong · v1 (2026-06-24)",Font(italic=True,size=10))
gline(5,"1 · MỤC ĐÍCH / PURPOSE",B,LBLUE)
gline(6,"• Thư viện chuẩn hóa các hạng mục kiểm IPQC theo dòng sản phẩm, dạng 1 dòng/1 check-item — sẵn sàng import vào app CMES.")
gline(7,"• Standardized IPQC check-item library by product line, one row per item — ready to import into CMES.")
gline(8,"• Nguồn: checklist 4 nhóm đặc tính + Pareto lỗi thực tế IPQC 2025–2026 (9.8M nhãn, 1.7M lụa, 1.0M dập-CNC).")
gline(10,"2 · CÁCH DÙNG SHEET / SHEETS",B,LBLUE)
gline(11,"• LIBRARY — bảng cái: mỗi dòng = 1 check-item. Lọc cột 'Dòng SP' để xem từng dòng (LABEL / SILK / PRESS_CNC).")
gline(12,"• DEFECT_CODES — danh mục mã lỗi (NG reason) theo Pareto, dùng làm dropdown lý do NG trong app.")
gline(13,"• AQL — bảng tra cỡ mẫu ISO 2859-1 (Single, Normal, Level II).")
gline(15,"3 · QUY ƯỚC CỘT / KEY COLUMNS",B,LBLUE)
gline(16,"• ItemID: khóa ổn định (vd LBL-A02) — dùng làm key import, KHÔNG đổi.")
gline(17,"• Mức/Severity: ◆ Critical (NG=DỪNG) · ● Major (sửa trước khi chạy) · ○ Minor (ghi nhận, sửa trong ca).")
gline(18,"• AQL: ◆→0.65 · ●→1.5 · ○→4.0. Cỡ mẫu = FAI 100% mẫu đầu + AQL theo bảng (sheet AQL).")
gline(19,"• Short-form = Y: hạng mục cốt lõi (Critical + vital-few Pareto + test chức năng FAI) → app hiện form ngắn mặc định; mở rộng xem đủ.")
gline(20,"• Loại KT: Visual (soi) · Measure (đo) · Functional (test phá hủy, mẫu nhỏ + định kỳ/ca).")
gline(21,"• Điều kiện áp dụng: hạng mục chỉ áp dụng khi có (vd 'Có keo', 'Có sơn/in') → app cho phép N/A.")
gline(23,"4 · KẾT QUẢ / RESULT",B,LBLUE)
gline(24,"• Mỗi item: OK · NG · N/A. Nếu NG → chọn Mã lỗi (DEFECT_CODES) + ghi chú. Bất kỳ ◆ = NG → lô DỪNG, không chạy SX.")
gline(25,"• KẾT LUẬN lô tự tính: có ◆NG → REJECT; còn lại theo Ac/Re của AQL.")
gline(27,"5 · NGUỒN / SOURCE",B,LBLUE)
gline(28,"Test chức năng theo ISO 2409 / ASTM D3359 (cross-cut), tape, IPA rub, ISO 15415/16 (barcode). Lấy mẫu ISO 2859-1.")

# ===================== LIBRARY =====================
ws = wb.create_sheet("LIBRARY")
ws.sheet_view.showGridLines=False
headers = ["ItemID","Dòng SP\nLine","Nhóm\nGroup","Mã\nCode","Hạng mục (VI)","Item (EN)",
           "Tiêu chuẩn chấp nhận (VI)","Acceptance (EN)","Phương pháp · Dụng cụ","Mức\nSeverity",
           "AQL","Cỡ mẫu / Sampling","Loại KT","Mã lỗi\nDefect","% Pareto","Short\nform",
           "ISO ref","Điều kiện áp dụng","Ghi chú / Note"]
widths = [10,9,13,6,30,22,34,24,30,10,6,20,11,10,8,7,18,16,26]
for i,(h,w) in enumerate(zip(headers,widths),1):
    c=ws.cell(row=1,column=i,value=h); c.font=H; c.fill=PatternFill("solid",fgColor=NAVY)
    c.alignment=ctr; c.border=border
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[1].height=30
ws.freeze_panes="E2"

r=2
for pref,lname,rows in LINES:
    for it in rows:
        code,grp,nvi,nen,avi,aen,meth,sev,insp,dfc,par,short,iso,cond,note = it
        sampling = "FAI 100% + AQL "+AQL[sev] if insp!="Functional" else "FAI 3–5pc + định kỳ/ca (phá hủy)"
        vals=[f"{pref}-{code}",pref if pref!="LBL" else "LABEL", grp, code, nvi, nen, avi, aen, meth,
              f"{SYM[sev]} {sev}", AQL[sev], sampling, insp, dfc, (par if par else ""),
              "Y" if short else "", iso, cond, note]
        vals[1] = {"LBL":"LABEL","DGT":"DIGITAL","SLK":"SILK","PNC":"PRESS_CNC"}[pref]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v); c.font=N; c.border=border
            c.alignment = ctr if ci in (2,3,4,10,11,13,14,15,16) else wrap
        # severity color
        sc=ws.cell(row=r,column=10); sc.font=Font(name="Calibri",size=9,bold=True,color={"Critical":CR,"Major":MA,"Minor":MI}[sev])
        # short-form highlight row light
        if short:
            for ci in range(1,20):
                ws.cell(row=r,column=ci).fill=PatternFill("solid",fgColor=VITAL)
        r+=1
ws.auto_filter.ref=f"A1:S{r-1}"

# ===================== DEFECT_CODES =====================
ws = wb.create_sheet("DEFECT_CODES")
ws.sheet_view.showGridLines=False
dh=["Dòng SP / Line","Mã lỗi / Code","Tên lỗi / Defect","Số lượt NG","% lỗi","Lũy kế %","Vital few (80%)"]
dw=[14,12,34,12,9,10,14]
for i,(h,w) in enumerate(zip(dh,dw),1):
    c=ws.cell(row=1,column=i,value=h);c.font=H;c.fill=PatternFill("solid",fgColor=BLUE);c.alignment=ctr;c.border=border
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[1].height=24; ws.freeze_panes="A2"
r=2
lname={"LBL":"LABEL","DGT":"DIGITAL","SLK":"SILK","PNC":"PRESS_CNC"}
for pref in ["LBL","DGT","SLK","PNC"]:
    cum=0.0
    for code,name,cnt,pct in DEFECTS[pref]:
        cum = cum+(pct or 0)
        vf = "★" if (pct and cum<=80.5 and pct>0) else ""
        vals=[lname[pref],code,name,(cnt if cnt else ""),(f"{pct}%" if pct is not None else "—"),
              (f"{cum:.1f}%" if pct is not None else "—"), vf]
        for ci,v in enumerate(vals,1):
            c=ws.cell(row=r,column=ci,value=v);c.font=N;c.border=border
            c.alignment=ctr if ci in (1,2,4,5,6,7) else wrap
        if vf:
            for ci in range(1,8): ws.cell(row=r,column=ci).fill=PatternFill("solid",fgColor=VITAL)
        r+=1
    r+=1  # blank sep

# ===================== AQL =====================
ws = wb.create_sheet("AQL")
ws.sheet_view.showGridLines=False
ws.cell(row=1,column=1,value="TẦN SUẤT & CỠ MẪU — AQL ISO 2859-1 / ANSI Z1.4 (Single, Normal, Level II)").font=B
ws.cell(row=2,column=1,value="FAI/FPI = 100% mẫu đầu (3–5pc). Bảng = lấy mẫu in-process. Ac=accept, Re=reject. ◆→0.65 ●→1.5 ○→4.0.").font=N
ah=["Cỡ lô / Lot size","Code","Cỡ mẫu n","AQL 0.65 ◆","AQL 1.5 ●","AQL 4.0 ○"]
for i,h in enumerate(ah,1):
    c=ws.cell(row=4,column=i,value=h);c.font=H;c.fill=PatternFill("solid",fgColor=BLUE);c.alignment=ctr;c.border=border
    ws.column_dimensions[get_column_letter(i)].width=16
AQLT=[("2–8","A",2,"↓","↓","↓"),("9–15","B",3,"↓","↓","↓"),("16–25","C",5,"↓","↓","↓"),
("26–50","D",8,"↓","↓","0 / 1"),("51–90","E",13,"↓","↓","1 / 2"),("91–150","F",20,"↓","0 / 1","2 / 3"),
("151–280","G",32,"0 / 1","1 / 2","3 / 4"),("281–500","H",50,"1 / 2","2 / 3","5 / 6"),
("501–1200","J",80,"1 / 2","3 / 4","7 / 8"),("1201–3200","K",125,"2 / 3","5 / 6","10 / 11"),
("3201–10000","L",200,"3 / 4","7 / 8","14 / 15"),("10001–35000","M",315,"5 / 6","10 / 11","21 / 22"),
("35001–150000","N",500,"7 / 8","14 / 15","↑"),("150001–500000","P",800,"10 / 11","21 / 22","↑"),
("≥ 500001","Q",1250,"14 / 15","↑","↑")]
r=5
for row in AQLT:
    for ci,v in enumerate(row,1):
        c=ws.cell(row=r,column=ci,value=v);c.font=N;c.border=border;c.alignment=ctr
    r+=1
ws.cell(row=r+1,column=1,value="Quy tắc: lỗi ≤ Ac → NHẬN lô; ≥ Re → LOẠI. ↓/↑ = chuyển plan cỡ lô liền kề. Test chức năng (D) phá hủy → mẫu nhỏ theo spec.").font=N

out="IPQC_Library_CMES_v2.xlsx"
wb.save(out)
print("SAVED:",out)
print("LIBRARY rows:", sum(len(r) for _,_,r in LINES))
for pref,lname,rows in LINES:
    sc=sum(1 for x in rows if x[11])
    print(f"  {lname}: {len(rows)} items, short-form {sc}")
